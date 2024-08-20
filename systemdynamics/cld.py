import os
import pandas as pd
import numpy as np
import os
import datetime
import json
from types import SimpleNamespace
from openpyxl import load_workbook

class Extract:
    def __init__(self, setting_name):
        current_path = os.getcwd()
        folders = current_path.split('/')
        if folders.count("systemdynamics") > 1:
            current_path = "/".join(current_path.split('/')[:-1])
        self.current_path = current_path
        file_path = os.path.join(current_path, f"{setting_name}.xlsx")
        self.file_path = file_path
        settings_path = os.path.join(current_path, f'{setting_name}.json')
        self.setting_name = setting_name
        self.settings_path = settings_path
        self.variables = []
        self.var_to_type = {}
        self.adjacency_matrix = None
        self.interactions_matrix = None
        self.get_settings() # Get settings from the json file
        self.test_extraction()  # Call the test_extraction function when the class is loaded

    def get_settings(self):
        """ Get the settings from the json file
        """
        with open(self.settings_path) as f:
            settings = json.load(f)
        s = SimpleNamespace(**settings)
        s.setting_name = self.setting_name
        curr_time = (str(datetime.datetime.now())[0:10])  # Create a new folder for each date

        if s.save_results:  # Create a directory to store results
            folder_path = os.path.join(self.current_path,"Results", f"{curr_time}_{self.setting_name}")
            
            if not os.path.exists(folder_path):
                os.mkdir(folder_path)

            with open(os.path.join(folder_path, f"used_settings_{self.setting_name}.json"), 'w+') as f:
                json.dump(settings, f, indent=2)  # Store current settings

            s.save_path = os.path.join("Results", curr_time + '_' + self.setting_name + "/")  # Path for saving the results
        self.s = s

    def extract_settings(self):
        """ Extract all settings based on the json and Kumu files
        """
        s = self.s
        # Load the adjacency matrix from the KUmu file
        variable_names, var_to_type_init, adjacency_matrix, interactions_matrix = self.adjacency_matrix_from_kumu()  

        if s.interaction_terms:
            if np.abs(interactions_matrix).sum() > 0:
                #s.interaction_terms = True
                print("Solving an SDM with interaction terms.")
                if s.solve_analytically and s.interaction_terms:
                    print("Cannot solve analytically with interaction terms so will proceed with numerical solution.")
            else:
                print("No interaction terms specified so will solve linear SDM.")
                s.interaction_terms = False
    
        # Load variable names and automatically fill any spaces with underscores
        s.stocks = [var.replace(" ", "_") for var in variable_names if var_to_type_init[var] == 'stock']
        s.auxiliaries = [var.replace(" ", "_") for var in variable_names if var_to_type_init[var] == 'auxiliary']
        s.constants = [var.replace(" ", "_") for var in variable_names if var_to_type_init[var] == 'constant']
        s.variables = [var.replace(" ", "_") for var in variable_names]  # s.auxiliaries + s.stocks + s.constants
        s.stocks_and_constants = [var.replace(" ", "_") for var in variable_names if var_to_type_init[var] in ['stock', 'constant']]
        s.stocks_and_auxiliaries = [var.replace(" ", "_") for var in variable_names if var_to_type_init[var] in ['stock', 'auxiliary']]
        s.var_to_type = {var.replace(" ", "_") : var_to_type_init[var] for var in variable_names}
        s.variable_of_interest = "_".join(s.variable_of_interest.split(" "))  # Ensure the variable of interest is formulated with underscores
        s.simulate_interventions = True  # Always simulate interventions
        
        # Create dataframe with adjacency matrix
        df_adj = pd.DataFrame(adjacency_matrix,
                            columns=s.variables, index=s.variables) 

        np.random.seed(s.seed)  # Set seed for reproducibility

        ### Check if any constants have incoming links
        for const in s.constants:
            num_links = np.sum(np.abs(df_adj.loc[const, :]))
            if num_links != 0:
                #if s.remove_incoming_links_constants:
                print(f'Removed {num_links} incoming links for constant {const}')
                df_adj.loc[const, :] = 0
                #else:
                #    raise(Exception(f'Number of incoming links for constant {const} is {num_links}, should be zero.'))

        # Set the SDM simulation timesteps to store 
        s.t_eval = np.array(np.array([0.0] + list(np.linspace(0, s.t_end,
                                                            int(s.t_end/s.dt) + 1)[1:])))

        # If solving the system numerically, set the solver
        s.solver = 'LSODA'  # 'LSODA' automatically switches between stiff and non-stiff methods since stiffness is not always known.

        # Select variables to simulated interventions for; all variables except the var of interest by default
        s.intervention_variables = [var for var in s.variables if var != s.variable_of_interest]  

        # If double factor interventions selected, add double factor interventions 
        if s.double_factor_interventions and s.interaction_terms == False:
            print("Without interaction terms, double factor interventions are not meaningful. Setting double_factor_interventions to False.")
            s.double_factor_interventions = 0

        if s.double_factor_interventions:
            double_intervention_variables = []
            for i, var in enumerate(s.intervention_variables):
                for j in range(i + 1, len(s.intervention_variables)):
                    var_2 = s.intervention_variables[j]
                    double_intervention_variables += [var + '+' + var_2]
            
            s.intervention_variables += double_intervention_variables

        s.df_adj = df_adj  # Save the adjacency matrix to the settings
        s.interactions_matrix = interactions_matrix # Save the interactions matrix to the settings

        # Add the interactions to the adjacency matrix for the identification of feedback loops with interaction terms
        s.df_adj_incl_interactions = s.df_adj.copy()
        to_list, from1_list, from2_list = np.nonzero(s.interactions_matrix)
        for i in range(int(np.abs(s.interactions_matrix).sum())):
            to, from1, from2 = to_list[i], from1_list[i], from2_list[i]
            value = s.interactions_matrix[to, from1, from2]
            # Ensure that the interaction is nonzero in the adjacency matrix
            s.df_adj_incl_interactions.loc[s.df_adj_incl_interactions.index[to],
                                           s.df_adj_incl_interactions.columns[from1]] = value
            s.df_adj_incl_interactions.loc[s.df_adj_incl_interactions.index[to],
                                           s.df_adj_incl_interactions.columns[from2]] = value
            
        self.s = s  # Save the settings

        return s

    def extract_adjacency_matrix(self):
        """Extract the adjacency matrix from an Excel table exported from Kumu (Kumu.io).
        The Kumu excel file contains one sheet with the CLD's variables ('Elements').
        It also contains a sheet with the CLD's causal links ('Connections').
        If there are known interactions in the system, these can be added in the 'Interactions' sheet.
        """
        # Read the elements, connections, and interactions sheets in the Kumu Excel file
        df_e = pd.read_excel(self.file_path, sheet_name="Elements")
        df_c = pd.read_excel(self.file_path, sheet_name="Connections")

        # Extract relevant columns
        df_e = df_e[["Label", "Type"]]   
        df_c = df_c[["From", "Type", "To"]]

        # Extract variables from the Elements 
        self.variables = list(df_e["Label"])
        self.var_to_type = dict(zip(list(df_e["Label"]), list(df_e["Type"])))

        # Create an empty adjacency matrix
        num_variables = len(self.variables)
        self.adjacency_matrix = np.zeros((num_variables, num_variables))

        # Populate the adjacency matrix
        for i, origin in enumerate(df_c["From"]):
            destination = df_c["To"][i]

            # Determine the polarity
            polarity = 0
            temp = df_c["Type"][i]
            if str(temp) == '+':
                polarity = 1
            elif str(temp) == '-':
                polarity = -1

            # Calculate indices
            origin_index = self.variables.index(origin)
            destination_index = self.variables.index(destination)

            # Add polarity to adjacency matrix
            self.adjacency_matrix[destination_index, origin_index] = polarity

    def extract_interactions_matrix(self):
        """Extract the interactions matrix from the 'Interactions' sheet in the Kumu Excel file."""
        wb = load_workbook(self.file_path, read_only=True)   # open an Excel file and return a workbook

        if 'Interactions' in wb.sheetnames:
            df_i = pd.read_excel(self.file_path, sheet_name="Interactions")
            df_i = df_i[["From1", "From2", "Type", "To"]]

            # Create an empty matrix to annotate interactions
            num_variables = len(self.variables)
            self.interactions_matrix = np.zeros((num_variables, num_variables, num_variables))

            # Populate the interactions matrix
            for i, origin_1 in enumerate(df_i["From1"]):
                origin_2 = df_i["From2"][i]
                destination = df_i["To"][i]

                # Determine the polarity
                polarity = 0
                temp = df_i["Type"][i]
                if str(temp) == '+':
                    polarity = 1
                elif str(temp) == '-':
                    polarity = -1

                # Calculate indices
                origin_1_index = self.variables.index(origin_1)
                origin_2_index = self.variables.index(origin_2)
                destination_index = self.variables.index(destination)

                # Add polarity to interactions matrix
                self.interactions_matrix[destination_index, origin_2_index, origin_1_index] = polarity
        

    def adjacency_matrix_from_kumu(self):
        """Run the CLD analysis by extracting the adjacency matrix and interactions matrix."""
        self.extract_adjacency_matrix()
        self.extract_interactions_matrix()
        return self.variables, self.var_to_type, self.adjacency_matrix, self.interactions_matrix


### TESTING ###
    def test_extraction(self):
        """Test the CLD extraction by creating an examplar Kumu table and comparing the results."""
        # Create a sample evidence table
        data = {
            "From": ["A", "B", "C"],
            "Type": ["+", "-", "+"],
            "To": ["B", "C", "A"]
        }

        data_int = {
            "From1": ["A", "B"],
            "From2": ["C", "C"],
            "Type": ["+", "+"],
            "To": ["B", "A"]
        }

        df_e = pd.DataFrame(data["From"], columns=["Label"])
        df_e["Type"] = ["stock", "auxiliary", "constant"]
        df_c = pd.DataFrame(data)
        df_i = pd.DataFrame(data_int)

        # Save the evidence table to an Excel file
        original_file_path = self.file_path
        test_file_path = os.path.join(os.path.dirname(__file__), '..', 'test_files', 'evidence_table.xlsx')

        # Ensure the directory exists
        os.makedirs(os.path.dirname(test_file_path), exist_ok=True)

        with pd.ExcelWriter(test_file_path) as writer:
            df_e.to_excel(writer, sheet_name='Elements', index=False)
            df_c.to_excel(writer, sheet_name='Connections', index=False)
            df_i.to_excel(writer, sheet_name='Interactions', index=False)

        # Run the extraction
        self.file_path = test_file_path
        self.adjacency_matrix_from_kumu()
        self.file_path = original_file_path  # Set the original file path again
    
        # Define the expected results
        expected_adjacency_matrix = np.array([[0, 0, 1],
                                              [1, 0, 0],
                                              [0, -1, 0]])

        expected_interactions_matrix = np.array([[[0, 0, 0],
                                                  [0, 0, 0],
                                                  [0, 1, 0]], 
                                                 [[0, 0, 0],
                                                  [0, 0, 0],
                                                  [1, 0, 0]], 
                                                 [[0, 0, 0],
                                                  [0, 0, 0],
                                                  [0, 0, 0]]])

        # Assess the results
        assert np.all(expected_adjacency_matrix == self.adjacency_matrix)
        assert np.all(expected_interactions_matrix == self.interactions_matrix)
        assert np.all([x in self.variables for x in data["From"]])
        assert np.all([x in data["From"] for x in self.variables])
        print("Test for loading KUMU table passed.")
    